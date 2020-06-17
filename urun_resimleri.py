import mysql.connector
from bs4 import BeautifulSoup 
import requests
from openpyxl import Workbook

a=1
wb=Workbook()
ws=wb.active

mydb = mysql.connector.connect(
	host="10.0.0.101",
	user="root",
	password="q1w2e3r4t5")

mycursor = mydb.cursor()

mycursor.execute("use crawlers")
mycursor.execute("select distinct product_url from price")

result= mycursor.fetchall()

for i in range(len(result)):
	r = requests.get(result[i][0])
	source = BeautifulSoup(r.content,"lxml")
	ws.cell(column=1,row=i+1,value = result[i][0])
	ws.cell(column=2,row=i+1,value = source.findAll('img',alt="Product")[0]['src'])
	if i % 100 == 0:
		print(a)
		a+=1 
wb.save("deneme.xlsx")
